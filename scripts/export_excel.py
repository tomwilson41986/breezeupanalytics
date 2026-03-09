#!/usr/bin/env python3
"""
Export Excel Report — Ratings, Stride Data & Detailed Times per Hip
====================================================================
Merges under-tack catalog data with stride/rating data into a formatted
Excel workbook with three sheets:

  1. Ratings Overview   — hip, horse info, rating, rank, distance/time
  2. Stride Data        — detailed stride metrics (frequency, length, diff)
  3. Detailed Times     — per-hip timing breakdown (UT time, GO time, set, group)

Usage
-----
    python scripts/export_excel.py
    python scripts/export_excel.py --output output/obs_march_2026_report.xlsx
    python scripts/export_excel.py --rated data/ratings/obsMarch\ stride_rated.csv \
                                   --catalog data/under-tack/obs_march_2026/latest.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paths (defaults) ──────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_RATED = ROOT / "data" / "ratings" / "obsMarch stride_rated.csv"
DEFAULT_CATALOG = ROOT / "data" / "under-tack" / "obs_march_2026" / "latest.json"
DEFAULT_OUTPUT = ROOT / "output" / "OBS_March_2026_Report.xlsx"

# ── Style constants ───────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
RATING_HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RATING_MID_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RATING_LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def load_catalog(path: Path) -> pd.DataFrame:
    """Load the under-tack catalog JSON into a DataFrame."""
    with open(path) as f:
        data = json.load(f)
    hips = data.get("hips", data) if isinstance(data, dict) else data
    df = pd.DataFrame(hips)
    df.rename(columns={"hip_number": "Hip"}, inplace=True)
    return df


def load_rated(path: Path) -> pd.DataFrame:
    """Load the rated CSV."""
    text = path.read_text()
    sep = "\t" if "\t" in text.split("\n")[0] else ","
    df = pd.read_csv(path, sep=sep, on_bad_lines="warn")
    df.columns = df.columns.str.strip()
    # Drop unnamed/empty columns
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    return df


def merge_data(rated: pd.DataFrame, catalog: pd.DataFrame) -> pd.DataFrame:
    """Merge rated stride data with catalog horse info."""
    rated["Hip"] = pd.to_numeric(rated["Hip"], errors="coerce")
    catalog["Hip"] = pd.to_numeric(catalog["Hip"], errors="coerce")
    merged = rated.merge(catalog, on="Hip", how="left")
    return merged.sort_values("Hip").reset_index(drop=True)


def _apply_header_style(ws, ncols: int):
    """Style the header row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=10, max_width=30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _apply_body_style(ws, nrows: int, ncols: int):
    """Apply body formatting."""
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_rating_colors(ws, rating_col_idx: int, nrows: int):
    """Conditionally color rating cells: green >75, yellow 40-75, red <40."""
    for row in range(2, nrows + 2):
        cell = ws.cell(row=row, column=rating_col_idx)
        try:
            val = float(cell.value) if cell.value is not None else None
        except (ValueError, TypeError):
            continue
        if val is None:
            continue
        if val >= 75:
            cell.fill = RATING_HIGH_FILL
        elif val >= 40:
            cell.fill = RATING_MID_FILL
        else:
            cell.fill = RATING_LOW_FILL


def _distance_label(val) -> str:
    """Convert distance feet to a friendly label."""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return str(val) if pd.notna(val) else ""
    if abs(v - 205) < 10:
        return "1/8 mi (205 ft)"
    if abs(v - 404.32) < 10:
        return "1/4 mi (404 ft)"
    return f"{v:.0f} ft"


def write_ratings_sheet(ws, merged: pd.DataFrame):
    """Sheet 1: Ratings Overview — one row per hip with key info + rating."""
    cols = [
        ("Hip", "Hip"),
        ("Horse Name", "horse_name"),
        ("Sex", "sex"),
        ("Sire", "sire"),
        ("Dam", "dam"),
        ("Dam Sire", "dam_sire"),
        ("Consignor", "consignor"),
        ("State Bred", "state_bred"),
        ("Distance", "Distance UT"),
        ("Time UT (s)", "Time UT"),
        ("Time GO (s)", "Time GO"),
        ("Rating", "Rating"),
        ("Mean Rank", "Mean Rank"),
    ]

    # Headers
    for c, (header, _) in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=header)

    # Data rows
    for r, (_, row) in enumerate(merged.iterrows(), 2):
        for c, (_, src) in enumerate(cols, 1):
            val = row.get(src)
            if src == "Distance UT" and pd.notna(val):
                val = _distance_label(val)
            elif src == "Hip" and pd.notna(val):
                val = int(val)
            elif src in ("Rating", "Mean Rank") and pd.notna(val):
                val = round(float(val), 1)
            elif src in ("Time UT", "Time GO") and pd.notna(val):
                val = round(float(val), 2)
            elif pd.isna(val) if isinstance(val, float) else val is None:
                val = ""
            ws.cell(row=r, column=c, value=val)

    nrows = len(merged)
    _apply_header_style(ws, len(cols))
    _apply_body_style(ws, nrows, len(cols))
    _auto_width(ws)

    # Color-code the Rating column
    rating_idx = next(i for i, (h, _) in enumerate(cols, 1) if h == "Rating")
    _apply_rating_colors(ws, rating_idx, nrows)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def write_stride_sheet(ws, merged: pd.DataFrame):
    """Sheet 2: Stride Data — detailed stride metrics per hip."""
    cols = [
        ("Hip", "Hip"),
        ("Distance", "Distance UT"),
        ("Stride Freq UT (Hz)", "Stride Frequency UT"),
        ("Stride Freq GO (Hz)", "Stride Frequency GO"),
        ("Stride Length UT (m)", "Stride Length UT"),
        ("Stride Length GO (m)", "Stride Length GO"),
        ("Stride Length UT (ft)", "Stride Length UT (ft)"),
        ("Stride Length GO (ft)", "Stride Length GO (ft)"),
        ("Stride Diff (m)", "diff"),
        ("Rank Stride UT", "Rank Stride Length UT (ft)"),
        ("Rank Stride GO", "Rank Stride Length GO (ft)"),
        ("Rank Diff", "Rank diff"),
    ]

    for c, (header, _) in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=header)

    for r, (_, row) in enumerate(merged.iterrows(), 2):
        for c, (_, src) in enumerate(cols, 1):
            val = row.get(src)
            if src == "Distance UT" and pd.notna(val):
                val = _distance_label(val)
            elif src == "Hip" and pd.notna(val):
                val = int(val)
            elif pd.notna(val) if isinstance(val, (int, float)) else val is not None:
                try:
                    val = round(float(val), 2)
                except (ValueError, TypeError):
                    pass
            else:
                val = ""
            ws.cell(row=r, column=c, value=val)

    nrows = len(merged)
    _apply_header_style(ws, len(cols))
    _apply_body_style(ws, nrows, len(cols))
    _auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def write_times_sheet(ws, merged: pd.DataFrame):
    """Sheet 3: Detailed Times — per-hip timing breakdown."""
    cols = [
        ("Hip", "Hip"),
        ("Horse Name", "horse_name"),
        ("Distance", "Distance UT"),
        ("Time UT (s)", "Time UT"),
        ("Time GO (s)", "Time GO"),
        ("UT Date", "ut_actual_date"),
        ("Set", "ut_set"),
        ("Group", "ut_group"),
        ("Rank Time UT", "Rank Time UT"),
        ("Rank Time GO", "Rank Time GO"),
        ("Rating", "Rating"),
        ("Sire", "sire"),
        ("Consignor", "consignor"),
        ("Status", "in_out_status"),
    ]

    for c, (header, _) in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=header)

    for r, (_, row) in enumerate(merged.iterrows(), 2):
        for c, (_, src) in enumerate(cols, 1):
            val = row.get(src)
            if src == "Distance UT" and pd.notna(val):
                val = _distance_label(val)
            elif src == "Hip" and pd.notna(val):
                val = int(val)
            elif src in ("Time UT", "Time GO") and pd.notna(val):
                val = round(float(val), 2)
            elif src in ("Rating",) and pd.notna(val):
                val = round(float(val), 1)
            elif src in ("Rank Time UT", "Rank Time GO") and pd.notna(val):
                try:
                    val = int(float(val))
                except (ValueError, TypeError):
                    pass
            elif pd.isna(val) if isinstance(val, float) else val is None:
                val = ""
            ws.cell(row=r, column=c, value=val)

    nrows = len(merged)
    _apply_header_style(ws, len(cols))
    _apply_body_style(ws, nrows, len(cols))
    _auto_width(ws)

    # Color-code Rating column here too
    rating_idx = next(i for i, (h, _) in enumerate(cols, 1) if h == "Rating")
    _apply_rating_colors(ws, rating_idx, nrows)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def export_excel(rated_path: Path, catalog_path: Path, output_path: Path):
    """Main export: build and save the Excel workbook."""
    print(f"Loading rated data from {rated_path}")
    rated = load_rated(rated_path)

    print(f"Loading catalog data from {catalog_path}")
    catalog = load_catalog(catalog_path)

    merged = merge_data(rated, catalog)
    total_hips = len(merged)
    rated_count = merged["Rating"].notna().sum()
    print(f"Merged {total_hips} hips ({rated_count} with ratings)")

    wb = Workbook()

    # Sheet 1: Ratings Overview
    ws_ratings = wb.active
    ws_ratings.title = "Ratings Overview"
    write_ratings_sheet(ws_ratings, merged)

    # Sheet 2: Stride Data
    ws_stride = wb.create_sheet("Stride Data")
    write_stride_sheet(ws_stride, merged)

    # Sheet 3: Detailed Times
    ws_times = wb.create_sheet("Detailed Times")
    write_times_sheet(ws_times, merged)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nExcel report saved to {output_path}")
    print(f"  Sheets: Ratings Overview | Stride Data | Detailed Times")
    print(f"  Hips: {total_hips} total, {rated_count} rated")


def main(argv: list[str] | None = None):
    parser = argparse.ArgumentParser(
        description="Export Excel report with ratings, stride data & detailed times per hip"
    )
    parser.add_argument(
        "--rated", type=Path, default=DEFAULT_RATED,
        help="Path to rated CSV (default: data/ratings/obsMarch stride_rated.csv)",
    )
    parser.add_argument(
        "--catalog", type=Path, default=DEFAULT_CATALOG,
        help="Path to catalog JSON (default: data/under-tack/obs_march_2026/latest.json)",
    )
    parser.add_argument(
        "--output", "-o", type=Path, default=DEFAULT_OUTPUT,
        help="Output Excel path (default: output/OBS_March_2026_Report.xlsx)",
    )
    args = parser.parse_args(argv)

    if not args.rated.exists():
        print(f"Error: rated file not found: {args.rated}", file=sys.stderr)
        sys.exit(1)
    if not args.catalog.exists():
        print(f"Error: catalog file not found: {args.catalog}", file=sys.stderr)
        sys.exit(1)

    export_excel(args.rated, args.catalog, args.output)


if __name__ == "__main__":
    main()
